from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import InvalidSelectorException
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.edge.service import Service
from message import message_box
import os
import openpyxl as xl


excel_columns = {"folder_link": 2, "doc_link": 1, "result": 3}
excel_filename = 'links.xlsx'

os.system("cls")
script_dir = os.path.abspath(os.path.dirname( __file__ ))

print(f'Opening the Excel-file...')
wb_fullpath = script_dir + "\\" + excel_filename
try:
    wb = xl.load_workbook(wb_fullpath)
except Exception as e:
    print(f'Cannot open the excel file: {str(e)}, quitting...')
    message_box('Error', f'Cannot open the excel file: {str(e)}', 0)    
    quit()
print(f'{excel_filename} has been opened successfully')

print(f'Starting Edge...')
options = EdgeOptions()
options.add_argument("start-maximized")
options.add_experimental_option('excludeSwitches', ['enable-logging'])
service_path = script_dir + "\\msedgedriver.exe"
edgeBrowser = webdriver.Edge(service=Service(service_path), options=options)

try:
    edgeBrowser.get('http://sww-edw.sakhalinenergy.ru/aha_seic_sww/asp/default.asp?AHAContextID=1')
except Exception as e:
    print(f"{str(e)}, aborting the program")
    edgeBrowser.quit()
    quit()

ws = wb['Input Data']

# clear prev. data
print(f'Clearing data in {excel_filename}...')
for row in ws['A2':'C100']:
  for cell in row:
    cell.value = None
    cell.hyperlink = None

# ****************** SEARCH CRITERIA ***********************
doc_desc_criteria = "*LMR*"
# doc_desc_criteria = "*MARSHALLING*LZR*6101*"
# doc_desc_criteria = "*SYSTEM*LZP*"
doc_number_criteria = "6000*"
# **********************************************************

print(f'Requesting AHA for search by criteria...')
llink = "http://sww-edw.sakhalinenergy.ru/aha_seic_sww/asp/treeview/tree.asp?Option=ObjectSearch&" \
        "obj_type_id=7&" \
        "obj_type_name=Document&" \
        "cls_obj_name=AhaQryStdRel.FindObjectWoRev%28%27Document%27%2C%27document+issue+date%27%29&" \
        f"obj_name={doc_number_criteria}&obj_desc={doc_desc_criteria}"

edgeBrowser.get(llink)
edgeBrowser.execute_script("window.scrollTo(0, document.body.scrollHeight);")

element_xpath = f"//a[@class='object_link']"
try:
    elem_list = edgeBrowser.find_elements(By.XPATH, element_xpath)
except NoSuchElementException:
    message_box('Error', 'NoSuchElementException', 0)

print(f'Forming a doc list...') 
doc_list = list()
for e in elem_list:
    doc_list.append({"node_id": e.get_attribute(name='id'), "doc_number": e.get_attribute('innerText')})

print(f'{len(doc_list)} document has(ve) been found, processing...') 

row = 2

for doc in doc_list:
    node_id = doc["node_id"]
    doc_number = doc["doc_number"]
    print(f'Processing {doc_number} with {node_id=}...') 
    llink = f"http://sww-edw.sakhalinenergy.ru/aha_seic_sww/asp/relationsandmethods.asp?TreeNodeID={node_id}&ScrollPosX=0&ScrollPosY=0"
    
    edgeBrowser.get(llink)

    element_xpath = "//a[text()='Jump in Unica compound document']"
    try:
        llink = edgeBrowser.find_element(By.XPATH, element_xpath).get_attribute("href")
    except NoSuchElementException:
        ws.cell(row, excel_columns[
            "result"]).value = "Node with text equal to 'Jump in Unica compound document' is not found"
        continue
    except InvalidSelectorException:
        ws.cell(row, excel_columns["result"]).value = f'Invalid XPATH={element_xpath} expression'
        continue

    edgeBrowser.get(llink)

    window_after = edgeBrowser.window_handles[1]
    edgeBrowser.switch_to.window(window_after)

    folder_link = edgeBrowser.current_url
    ws.cell(row, excel_columns["folder_link"]).hyperlink = folder_link
    ws.cell(row, excel_columns["folder_link"]).value = doc_number
    ws.cell(row, excel_columns["folder_link"]).style = "Hyperlink"

    element_xpath = "//a[@data-otname='itemContainer']"
    try:
        doc_list = edgeBrowser.find_elements(By.XPATH, element_xpath)
    except NoSuchElementException:
        ws.cell(row, excel_columns[
            "result"]).value = f"Node with href={element_xpath} has not been found"
        continue
    except InvalidSelectorException:
        ws.cell(row, excel_columns["result"]).value = f'Invalid XPATH={element_xpath} expression'
        continue
    
    for doc in doc_list:
        doc_link = doc.get_attribute("href")
        ws.cell(row, excel_columns["doc_link"]).hyperlink = doc_link
        ws.cell(row, excel_columns["doc_link"]).value = doc.get_attribute("innerText")
        ws.cell(row, excel_columns["doc_link"]).style = "Hyperlink"
        row += 1
    
print(f'Saving the Excel-file...')
try:
  wb.save(wb_fullpath)
except Exception as e:
    print(f'Cannot save the excel file: {str(e)}')

wb.close()
message_box('End of script', 'Ok!', 0)
print(f"End of script")
# os.system('pause')
edgeBrowser.quit()
