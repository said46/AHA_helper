from selenium import webdriver
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.edge.service import Service

from message import message_box
import openpyxl as xl
import os

def save_excel_file(fn: str) -> None:
    try:
        wb.save(fn)
    except Exception as e:
        print(f'Cannot save the excel file: {str(e)}')


os.system("cls")
script_dir = os.path.abspath(os.path.dirname( __file__ ))
excel_filename = 'links.xlsx'
wb_fullpath = script_dir + "\\" + excel_filename

pdfs_folder_path = os.getcwd() + '\\_get_pdfs'

options = EdgeOptions()
options.add_experimental_option('prefs', {
                                            "download.prompt_for_download": False,
                                            "plugins.always_open_pdf_externally": True,
                                            "download.default_directory": pdfs_folder_path, 
                                            "download.directory_upgrade": True
                                         })
options.add_experimental_option('excludeSwitches', ['enable-logging'])

service_path = script_dir + "\\msedgedriver.exe"
print(f"Starting Edge...")
edgeBrowser = webdriver.Edge(service=Service(service_path), options=options)
edgeBrowser.maximize_window()

print(f"Opening {excel_filename}...")
wb = xl.load_workbook(wb_fullpath)
sheet = wb['Input Data']

# for each row
for count, row in enumerate(range(2, sheet.max_row + 1)):
    print(f"Processig row number {row} in {excel_filename}...")
    # if an empty row - stop
    if sheet.cell(row, 1).value is None:
        break        

    if sheet.cell(row, 3).value == "downloaded":
        continue
    
    if count == 55555:
        edgeBrowser.quit()
        save_excel_file(wb_fullpath)
        wb.close()
        quit()

    try:
        llink = sheet.cell(row, 1).hyperlink.target
    except Exception as e:
        sheet.cell(row, 2).value = f"{str(e)}" 
        save_excel_file(wb_fullpath)
        continue
    
    doc_number = sheet.cell(row, 1).value
    try:
        edgeBrowser.get(llink)
    except Exception as e:
        sheet.cell(row, 2).value = f"{str(e)}"        
        save_excel_file(wb_fullpath)
        continue

    sheet.cell(row, 3).value = f'downloaded'
    save_excel_file(wb_fullpath)

message_box('End of script', 'Ok!', 0)
print(f"End of script")
# input("Press Enter to continue...")

edgeBrowser.quit()
wb.close()
